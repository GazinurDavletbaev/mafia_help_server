import io
from fastapi import FastAPI, Response, HTTPException
from pydantic import BaseModel
from typing import List, Dict, Any
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side
import os
import re

app = FastAPI(title="Mafia Protocol API (Excel)")

# ========== МОДЕЛИ ==========
class PlayerData(BaseModel):
    seat: int
    name: str
    role: str
    fouls: int
    points: int
    bonus: float

class ProtocolData(BaseModel):
    tournament: str
    stage: str
    table: str
    game: str
    date: str
    time: str  # ← время игры
    judge: str
    bestMove: str
    protest: str
    winner: str
    players: List[PlayerData]
    nightActions: List[int]
    voteHistory: Dict[str, Dict[str, Any]]  # новый формат
    notes: List[str] = []          # ← ДОБАВИТЬ
    protestComment: str = ""       # ← ДОБАВИТЬ

# ========== ВСПОМОГАТЕЛЬНЫЕ ==========
def get_role_short(role: str) -> str:
    roles = {'don': 'Д', 'mafia': 'Ч', 'sheriff': 'Ш', 'citizen': 'К'}
    return roles.get(role, '?')


def parse_best_move(best_move_str: str) -> List[str]:
    if not best_move_str:
        return ['', '', '']
    parts = re.split(r'[,，\s]+', best_move_str.strip())
    result = [p.strip() for p in parts[:3]]
    while len(result) < 3:
        result.append('')
    return result


# ========== ОСНОВНОЙ ЭНДПОИНТ ==========
@app.post("/generate-protocol-excel")
async def generate_protocol_excel(data: ProtocolData):
    try:
        print("=" * 60)
        print("📥 ПОЛУЧЕН ЗАПРОС")
        print("=" * 60)
        
        print(f"📊 tournament: {data.tournament}")
        print(f"📊 stage: {data.stage}")
        print(f"📊 table: {data.table}")
        print(f"📊 game: {data.game}")
        print(f"📊 date: {data.date}")
        print(f"📊 time: {data.time}")
        print(f"📊 judge: {data.judge}")
        print(f"📊 bestMove: {data.bestMove}")
        print(f"📊 protest: {data.protest}")
        print(f"📊 winner: {data.winner}")
        print(f"📊 players: {len(data.players)} игроков")
        print(f"📊 nightActions: {data.nightActions}")
        print(f"📊 voteHistory: {data.voteHistory}")
        print("-" * 60)

        wb = load_workbook("blank_protocol2.xlsx")
        ws = wb.active

        # ========== 1. ШАПКА ==========
        ws['O4'] = data.tournament if data.tournament else ""
        ws['BA4'] = data.stage if data.stage else ""
        ws['K7'] = data.date if data.date else ""
        ws['V7'] = data.time if data.time else "00:00 — 00:00"  # ← время
        ws['AQ7'] = data.table if data.table else ""
        ws['BE7'] = data.game if data.game else ""

        # ========== 2. ТАБЛИЦА ИГРОКОВ ==========
        player_rows = [14, 18, 22, 26, 30, 34, 38, 42, 46, 50]

        for i, player in enumerate(data.players):
            if i >= 10:
                break
            row = player_rows[i]
            ws[f'F{row}'] = player.name  # ← F вместо E
            ws[f'AG{row}'] = get_role_short(player.role)
            ws[f'AN{row}'] = player.fouls
            ws[f'AU{row}'] = player.points
            ws[f'BB{row}'] = player.bonus

        # ========== 3. ИНФОРМАЦИЯ ==========
        winner_text = "КРАСНЫЕ" if data.winner == 'red' else 'ЧЁРНЫЕ'
        ws['AG55'] = winner_text

        best_parts = parse_best_move(data.bestMove)
        ws['Y59'] = best_parts[0] if len(best_parts) > 0 else ""
        ws['AC59'] = best_parts[1] if len(best_parts) > 1 else ""
        ws['AG59'] = best_parts[2] if len(best_parts) > 2 else ""

        best_move_player = data.nightActions[0] if data.nightActions and data.nightActions[0] > 0 else ""
        ws['BD59'] = best_move_player

        # Стрельба
        shoot_actions = data.nightActions[0::3]
        print(f"📊 shoot_actions (стрельбы): {shoot_actions}")
        shoot_cells = ['V63', 'Z63', 'AD63', 'AH63', 'AL63', 'AP63', 'AT63', 'AX63', 'BB63', 'BF63']
        for i, action in enumerate(shoot_actions):
            if i >= len(shoot_cells):
                break
            ws[shoot_cells[i]] = action if action > 0 else ""
            print(f"   📝 Стрельба {i+1}: {shoot_cells[i]} = {action if action > 0 else ''}")

        ws['T67'] = data.protest if data.protest else ""
        ws['Q71'] = data.judge if data.judge else ""

        # ========== 4. ГОЛОСОВАНИЯ (НОВЫЙ ФОРМАТ) ==========
        print("-" * 60)
        print("📊 ОБРАБОТКА ГОЛОСОВАНИЙ (НОВЫЙ ФОРМАТ)")
        print("-" * 60)

        if data.voteHistory:
            days = sorted([int(k) for k in data.voteHistory.keys()])
            print(f"📊 Дни: {days}")

            vote_idx = 0
            for day in days:
                if vote_idx >= 7:
                    break

                day_data = data.voteHistory[str(day)]
                rounds = day_data.get('rounds', [])
                eliminated = day_data.get('eliminated', False)
                elimination_votes = day_data.get('eliminationVotes', 0)
                result_players = day_data.get('result', [])

                print(f"\n📊 ДЕНЬ {day}: {len(rounds)} раундов")
                print(f"   result: {result_players}")
                print(f"   eliminationVotes: {elimination_votes}")

                # Берем все раунды для этого дня
                all_players = []
                all_votes = []
                all_revote_players = []
                all_revote_votes = []

                for round_idx, round_data in enumerate(rounds):
                    if round_idx == 0:
                        # Первый раунд — основное голосование
                        for player, count in round_data.items():
                            all_players.append(int(player))
                            all_votes.append(count)
                    else:
                        # Следующие раунды — переголосование
                        for player, count in round_data.items():
                            all_revote_players.append(int(player))
                            all_revote_votes.append(count)

                # Сортируем
                sorted_players = sorted(set(all_players))
                sorted_revote_players = sorted(set(all_revote_players))

                row_offset = vote_idx * 10
                row_players = 6 + row_offset
                row_votes = 8 + row_offset
                row_revote_players = 10 + row_offset
                row_revote_votes = 12 + row_offset
                row_result = 8 + row_offset

                slot_cols = ['BQ', 'BU', 'BY', 'CC', 'CG', 'CK', 'CO', 'CS', 'CW', 'DA']

                # Пишем основное голосование (игроки)
                for slot_idx, col in enumerate(slot_cols):
                    if slot_idx < len(sorted_players):
                        cell = f'{col}{row_players}'
                        ws[cell] = sorted_players[slot_idx]
                        print(f"   ✅ Игрок {sorted_players[slot_idx]} в {cell}")

                # Пишем основное голосование (голоса)
                for slot_idx, col in enumerate(slot_cols):
                    if slot_idx < len(sorted_players):
                        player = sorted_players[slot_idx]
                        # Находим голоса для этого игрока в первом раунде
                        vote_count = 0
                        if rounds and rounds[0]:
                            vote_count = rounds[0].get(str(player), 0)
                        cell = f'{col}{row_votes}'
                        ws[cell] = vote_count
                        print(f"   ✅ Голоса {vote_count} в {cell}")

                # ✅ Пишем переголосование (голоса) — если есть
                if sorted_revote_players and len(rounds) > 1:
                    # Раунд 2 — строки 10-11
                    revote_round_1 = rounds[1] if len(rounds) > 1 else {}
                    for slot_idx, col in enumerate(slot_cols):
                        if slot_idx < len(sorted_revote_players):
                            player = sorted_revote_players[slot_idx]
                            vote_count = revote_round_1.get(str(player), '-')
                            # ✅ Только левая верхняя ячейка (строка 10)
                            cell = f'{col}{row_revote_players}'
                            ws[cell] = vote_count
                            print(f"   ✅ Переголосование раунд 2: игрок {player} -> {vote_count} в {cell}")

                    # Раунд 3 — строки 12-13
                    if len(rounds) > 2:
                        revote_round_2 = rounds[2] if len(rounds) > 2 else {}
                        for slot_idx, col in enumerate(slot_cols):
                            if slot_idx < len(sorted_revote_players):
                                player = sorted_revote_players[slot_idx]
                                vote_count = revote_round_2.get(str(player), '-')
                                # ✅ Только левая верхняя ячейка (строка 12)
                                cell = f'{col}{row_revote_votes}'
                                ws[cell] = vote_count
                                print(f"   ✅ Переголосование раунд 3: игрок {player} -> {vote_count} в {cell}")

                # Пишем результат
                cell = f'DF{row_result}'
                if result_players:
                    ws[cell] = ', '.join(map(str, sorted(set(result_players))))
                    print(f"   ✅ Результат: {', '.join(map(str, sorted(set(result_players))))} в {cell}")
                else:
                    ws[cell] = '0'
                    print(f"   ✅ Результат: 0 в {cell}")

                vote_idx += 1
        else:
            print("📊 voteHistory пуст или None")

                # ========== 5. ПОЯСНЕНИЯ И КОММЕНТАРИИ ==========
        notes = data.notes if data.notes else []
        for i, row in enumerate(range(79, 93)):
            if i < len(notes) and notes[i]:
                ws[f'D{row}'] = notes[i]
                print(f"   ✅ Пояснение {i+1}: {notes[i]} в D{row}")
            else:
                ws[f'D{row}'] = ""

                protest_comment = data.protestComment if data.protestComment else ""
        if protest_comment:
            # Разбиваем текст по словам и переносим по 60 символов с учётом целых слов
            words = protest_comment.split()
            lines = []
            current_line = ""
            for word in words:
                if len(current_line) + len(word) + 1 <= 60:
                    if current_line:
                        current_line += " " + word
                    else:
                        current_line = word
                else:
                    if current_line:
                        lines.append(current_line)
                    current_line = word
            if current_line:
                lines.append(current_line)
            
            # Записываем строки в ячейки D94, D95, D96...
            for i, line in enumerate(lines[:10]):  # максимум 10 строк
                row = 94 + i
                ws[f'D{row}'] = line
                print(f"   ✅ Комментарий к протесту строка {i+1}: {line} в D{row}")
        # ========== 5. СОХРАНЕНИЕ ==========
        print("-" * 60)
        print("💾 СОХРАНЕНИЕ ФАЙЛА")
        print("-" * 60)
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        print("✅ Файл успешно создан")
        print("=" * 60)

        return Response(
            content=buffer.read(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename=protocol_{data.table}_{data.game}.xlsx",
                "Access-Control-Allow-Origin": "*"
            }
        )

    except Exception as e:
        print(f"❌ ОШИБКА: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/health")
async def health():
    return {"status": "ok"}


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8001)
