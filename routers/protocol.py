from fastapi import APIRouter, HTTPException, Response
from pydantic import BaseModel
from typing import List, Dict, Any
import io
import re
from openpyxl import load_workbook

router = APIRouter()

# ========== МОДЕЛИ ==========
class PlayerData(BaseModel):
    seat: int
    name: str
    role: str
    fouls: int
    points: int
    bonus: float
    rule: str = ""  # ✅ ДОБАВЛЕНО

class ProtocolData(BaseModel):
    tournament: str
    stage: str
    table: str
    game: str
    date: str
    time: str
    judge: str
    bestMove: str
    protest: str
    winner: str
    players: List[PlayerData]
    nightActions: List[int]
    voteHistory: Dict[str, Dict[str, Any]]
    notes: List[str] = []
    protestComment: str = ""

# ========== ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ ==========
def get_role_short(role: str) -> str:
    roles = {'don': 'Д', 'mafia': 'Ч', 'sheriff': 'Ш', 'citizen': 'К'}
    return roles.get(role, '?')

def parse_best_move(best_move_str: str) -> List[str]:
    if not best_move_str:
        return ['', '', '']
    parts = re.split(r'[,，\s]+', best_move_str.strip())
    result = [p.strip() for p in parts[:3]]
    while len(result) < 3:
        result.append('')
    return result

# ========== ЭНДПОИНТ ==========
@router.post("/generate")
async def generate_protocol(data: ProtocolData):
    try:
        print("=" * 60)
        print("📥 ПОЛУЧЕН ЗАПРОС НА ГЕНЕРАЦИЮ EXCEL")
        print("=" * 60)
        
        print(f"📊 tournament: {data.tournament}")
        print(f"📊 stage: {data.stage}")
        print(f"📊 table: {data.table}")
        print(f"📊 game: {data.game}")
        print(f"📊 date: {data.date}")
        print(f"📊 time: {data.time}")
        print(f"📊 judge: {data.judge}")
        print(f"📊 winner: {data.winner}")
        print(f"📊 players: {len(data.players)}")
        print("-" * 60)

        wb = load_workbook("blank_protocol2.xlsx")
        ws = wb.active

        # Шапка
        ws['O4'] = data.tournament if data.tournament else ""
        ws['BA4'] = data.stage if data.stage else ""
        ws['K7'] = data.date if data.date else ""
        ws['V7'] = data.time if data.time else "00:00 — 00:00"
        ws['AQ7'] = data.table if data.table else ""
        ws['BE7'] = data.game if data.game else ""

        # Игроки
        player_rows = [14, 18, 22, 26, 30, 34, 38, 42, 46, 50]
        for i, player in enumerate(data.players):
            if i >= 10:
                break
            row = player_rows[i]
            ws[f'F{row}'] = player.name
            ws[f'AG{row}'] = get_role_short(player.role)
            ws[f'AN{row}'] = player.fouls
            ws[f'AU{row}'] = player.points
            ws[f'BB{row}'] = player.bonus

        # Информация
        winner_text = "КРАСНЫЕ" if data.winner == 'red' else 'ЧЁРНЫЕ'
        ws['AG55'] = winner_text

        best_parts = parse_best_move(data.bestMove)
        ws['Y59'] = best_parts[0] if len(best_parts) > 0 else ""
        ws['AC59'] = best_parts[1] if len(best_parts) > 1 else ""
        ws['AG59'] = best_parts[2] if len(best_parts) > 2 else ""

        best_move_player = data.nightActions[0] if data.nightActions and data.nightActions[0] > 0 else ""
        ws['BD59'] = best_move_player

        # Стрельба
        shoot_actions = data.nightActions[0::3]
        shoot_cells = ['V63', 'Z63', 'AD63', 'AH63', 'AL63', 'AP63', 'AT63', 'AX63', 'BB63', 'BF63']
        for i, action in enumerate(shoot_actions):
            if i >= len(shoot_cells):
                break
            ws[shoot_cells[i]] = action if action > 0 else ""

        ws['T67'] = data.protest if data.protest else ""
        ws['Q71'] = data.judge if data.judge else ""

        # Голосования
        if data.voteHistory:
            days = sorted([int(k) for k in data.voteHistory.keys()])
            vote_idx = 0
            for day in days:
                if vote_idx >= 7:
                    break

                day_data = data.voteHistory[str(day)]
                rounds = day_data.get('rounds', [])
                result_players = day_data.get('result', [])

                row_offset = vote_idx * 10
                row_players = 6 + row_offset
                row_votes = 8 + row_offset
                row_revote_players = 10 + row_offset
                row_revote_votes = 12 + row_offset
                row_result = 8 + row_offset

                slot_cols = ['BQ', 'BU', 'BY', 'CC', 'CG', 'CK', 'CO', 'CS', 'CW', 'DA']

                if rounds:
                    round_1 = rounds[0] if len(rounds) > 0 else {}
                    sorted_players = sorted([int(k) for k in round_1.keys()])
                    for slot_idx, col in enumerate(slot_cols):
                        if slot_idx < len(sorted_players):
                            player = sorted_players[slot_idx]
                            ws[f'{col}{row_players}'] = player
                            ws[f'{col}{row_votes}'] = round_1.get(str(player), 0)

                    if len(rounds) > 1:
                        round_2 = rounds[1] if len(rounds) > 1 else {}
                        sorted_revote = sorted([int(k) for k in round_2.keys()])
                        for slot_idx, col in enumerate(slot_cols):
                            if slot_idx < len(sorted_revote):
                                player = sorted_revote[slot_idx]
                                ws[f'{col}{row_revote_players}'] = round_2.get(str(player), '-')

                    if len(rounds) > 2:
                        round_3 = rounds[2] if len(rounds) > 2 else {}
                        sorted_revote = sorted([int(k) for k in round_3.keys()])
                        for slot_idx, col in enumerate(slot_cols):
                            if slot_idx < len(sorted_revote):
                                player = sorted_revote[slot_idx]
                                ws[f'{col}{row_revote_votes}'] = round_3.get(str(player), '-')

                cell = f'DF{row_result}'
                if result_players:
                    ws[cell] = ', '.join(map(str, sorted(set(result_players))))
                else:
                    ws[cell] = '0'

                vote_idx += 1

        # Пояснения
        notes = data.notes if data.notes else []
        for i, row in enumerate(range(79, 93)):
            if i < len(notes) and notes[i]:
                ws[f'D{row}'] = notes[i]
            else:
                ws[f'D{row}'] = ""

        # Комментарий к протесту
        protest_comment = data.protestComment if data.protestComment else ""
        if protest_comment:
            words = protest_comment.split()
            lines = []
            current_line = ""
            for word in words:
                if len(current_line) + len(word) + 1 <= 60:
                    if current_line:
                        current_line += " " + word
                    else:
                        current_line = word
                else:
                    if current_line:
                        lines.append(current_line)
                    current_line = word
            if current_line:
                lines.append(current_line)
            
            for i, line in enumerate(lines[:10]):
                row = 94 + i
                ws[f'D{row}'] = line

        # Сохранение
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        print("✅ Excel успешно создан")
        print("=" * 60)

        return Response(
            content=buffer.read(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename=protocol_{data.table}_{data.game}.xlsx",
                "Access-Control-Allow-Origin": "*"
            }
        )

    except Exception as e:
        print(f"❌ ОШИБКА: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))